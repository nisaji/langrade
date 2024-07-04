import openpyxl
from langrade.infrastructure.document_retriever.parsers.base import BaseParser


class ExcelParser(BaseParser):
    def parse(self, file_path: str) -> str:
        workbook = openpyxl.load_workbook(file_path)
        content = []
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            content.append(f"Sheet: {sheet}")
            for row in worksheet.iter_rows(values_only=True):
                content.append(
                    "\t".join(str(cell) for cell in row if cell is not None)
                )  # noqa: E501
        return "\n".join(content)
